"""External-dataset upload parser (UI Slice 4 — the front door to the scratch write plane).

Transport-agnostic parsing for a user-uploaded CSV or XLSX file, producing the
exact ``{columns:[{name,type}], rows:[[...]]}`` shape that
``app.scratch_ingest.materialize`` already consumes — with ZERO change to the
write plane.  This is the *parse + column-mapping* half in front of the built,
D92-bound scratch materialize back-half (UI Slice 4, §2).

Design (why this is a NEW module, not an extension of ``admin_ingest``):
  ``admin_ingest`` is the warehouse bulk-load path (request-supplied CH creds,
  200 MB cap, schema-diff against a target table).  The upload front door is a
  session-scratch path (server creds, few-MB cap, no target table) — a different
  lifecycle and trust model.  But it IMPORTS and REUSES ``admin_ingest``'s
  proven, side-effect-free parser primitives (the same way ``scratch_ingest``
  reuses ``coerce``/``validate_identifier``):

    - ``infer_schema_streaming`` — (header, per-column inferred types, total)
    - ``_sanitize_columns``      — CSV header cells → unique CH identifiers
    - ``infer_column_type``      — per-column type inference for the XLSX path

  Cells stay STRINGS on both paths; the declared CH type is applied at insert by
  ``scratch_ingest._coerce_row`` → ``admin_ingest.coerce``.  So CSV and XLSX
  converge onto ONE coercion path, and no XLSX-native type reaches the write
  plane.

XLSX safety (UI Slice 4 §7 — the slice's main risk).  XLSX is a zip, so it
carries a decompression-bomb / zip-bomb risk plus openpyxl's XML entity history.
The mitigations here are:
  (1) the caller's pre-parse BYTE cap bounds the compressed input (endpoint layer);
  (2) ``read_only=True`` streams rows instead of loading the whole sheet;
  (3) ``data_only=True`` reads cached values and NEVER evaluates a formula;
  (4) the ROW cap is enforced *as rows are streamed* — the check runs before each
      row is appended, so a small compressed file that expands to millions of
      rows trips ``SCRATCH_TOO_LARGE`` after at most ``max_rows`` rows are held in
      memory, never after materializing the whole sheet;
  (5) every cell is stringified, so no XLSX-native object reaches the write plane.
"""

from __future__ import annotations

import csv
import io
from typing import Any

from app.admin_ingest import (
    _sanitize_columns,
    infer_column_type,
    infer_schema_streaming,
)
from app.scratch_ingest import ScratchTooLargeError

# ---------------------------------------------------------------------------
# Column-role → canonical join key (UI Slice 4 §1b, locked).
#
# The agent later JOINs the scratch table on these canonical keys, so the mapping
# renames a role-tagged column to exactly the key the agent expects.  A role
# other than "none" may appear at most once, and its rename target must not
# collide with a kept column's name (fail-closed — never silently drop/merge).
# ---------------------------------------------------------------------------
_ROLE_TO_KEY = {
    "EmployeeCode": "employee_code",
    "DepartmentCode": "department_code",
    "DepartmentName": "department_name",
}
_NONE_ROLES = frozenset({"none", "", None})

# Hard cap on the number of columns a single upload may declare (UI Slice 4 §7,
# resource-exhaustion hardening).  It closes TWO width-DoS vectors with one
# constant, checked BEFORE any per-column work:
#   - CSV: an 8 MB header of duplicate names (`a,a,a,…`) would drive
#     `_sanitize_columns`'s collision loop quadratic and allocate per-column
#     inference state for millions of columns.
#   - XLSX: a tiny compressed file can declare a huge sheet dimension (up to
#     openpyxl's 16,384 columns) so `read_only` pads every streamed row that
#     wide — a few-KB file expands to millions of stringified cells before the
#     ROW cap can trip.  We also pass `max_col=MAX_UPLOAD_COLUMNS + 1` to
#     `iter_rows`, so openpyxl never yields (or pads) past the cap regardless of
#     the declared dimension.
MAX_UPLOAD_COLUMNS = 512


# ---------------------------------------------------------------------------
# Exceptions
# ---------------------------------------------------------------------------


class UploadParseError(Exception):
    """The uploaded file could not be parsed.

    Raised for undecodable / unreadable input, an empty file, or a missing
    header.  The message is safe to surface (it never echoes row data).
    """

    def __init__(self, message: str) -> None:
        super().__init__(message)
        self.message = message


class UploadMappingError(Exception):
    """The column mapping was rejected — duplicate role or a rename collision.

    Rendered as ``400 UPLOAD_MAPPING_INVALID`` by the route.  Fail-closed; the
    upload never silently drops or merges a column.
    """

    def __init__(self, message: str) -> None:
        super().__init__(message)
        self.message = message


class UploadTooLargeError(Exception):
    """The uploaded body/file exceeds the front-door BYTE cap.

    Rendered as ``413 UPLOAD_TOO_LARGE`` by the route.  Distinct from the row-cap
    ``ScratchTooLargeError`` (SCRATCH_TOO_LARGE); this is the pre-parse byte guard.
    """

    def __init__(self, message: str) -> None:
        super().__init__(message)
        self.message = message


# ---------------------------------------------------------------------------
# Public surface
# ---------------------------------------------------------------------------


def parse_upload(
    content: bytes,
    filename: str,
    max_rows: int | None = None,
) -> tuple[list[dict[str, str]], list[list[Any]]]:
    """Dispatch CSV vs XLSX by extension; return ({name,type} columns, data rows).

    Columns are sanitized identifiers with inferred ClickHouse types — the exact
    ``{columns, rows}`` shape ``scratch_ingest.materialize`` consumes.  Cells stay
    STRINGS (coerced to their declared type at insert).

    ``max_rows`` (when not None) bounds the parse: a file whose data-row count
    exceeds it raises ``ScratchTooLargeError`` (``SCRATCH_TOO_LARGE``).  For XLSX
    this is enforced *as rows stream in*, so an over-cap sheet never fully
    materializes in memory (see the module docstring, mitigation #4).

    Raises ``UploadParseError`` on undecodable / unreadable / header-less input.
    """
    # DELIBERATE DEVIATION from §1a ("extension + content-type"): dispatch is
    # extension-only.  A mislabeled file (e.g. XLSX bytes named .csv) just takes
    # the CSV path and fails cleanly with UPLOAD_PARSE_ERROR (non-UTF-8) — content
    # sniffing would add complexity for no safety gain.
    if filename.lower().endswith(".xlsx"):
        return _parse_xlsx(content, max_rows)
    # Default to CSV for .csv and any other/unknown extension.
    return _parse_csv(content, max_rows)


def apply_mapping(
    columns: list[dict[str, str]],
    rows: list[list[Any]],
    mapping: dict[str, Any],
) -> tuple[list[dict[str, str]], list[list[Any]]]:
    """Rename role-tagged columns to their canonical join keys; validate the plan.

    ``mapping`` maps ``<sanitized_col_name> -> <role>`` where role is one of
    ``EmployeeCode`` | ``DepartmentCode`` | ``DepartmentName`` | ``none``.  A
    column absent from the mapping (or mapped to ``none``) keeps its sanitized
    name.  Keys that name no parsed column are ignored (the file is
    authoritative).  Rows are untouched — the rename is columns-only.

    Raises ``UploadMappingError`` if:
      - the mapping is not an object, or a value is not a recognized role;
      - a role (other than ``none``) is assigned to more than one column;
      - a rename target collides with a KEPT column's name.
    """
    if not isinstance(mapping, dict):
        raise UploadMappingError("'mapping' must be a JSON object.")

    existing_names = [c["name"] for c in columns]
    existing_set = set(existing_names)

    # Build the rename plan (old sanitized name -> canonical key), rejecting a
    # duplicate role along the way.
    renames: dict[str, str] = {}
    seen_roles: dict[str, str] = {}
    for col_name, role in mapping.items():
        if role in _NONE_ROLES:
            continue
        if role not in _ROLE_TO_KEY:
            raise UploadMappingError(f"Unknown column role {role!r}.")
        if col_name not in existing_set:
            # A mapping key that names no parsed column is ignored.
            continue
        if role in seen_roles:
            raise UploadMappingError(
                f"Role {role!r} is assigned to more than one column."
            )
        seen_roles[role] = col_name
        renames[col_name] = _ROLE_TO_KEY[role]

    # Collision check: a rename target must not equal any KEPT column's name.
    # (Two renames can never collide — roles are unique, so their targets are.)
    kept_names = {n for n in existing_names if n not in renames}
    for old_name, target in renames.items():
        if target in kept_names:
            raise UploadMappingError(
                f"Rename target {target!r} collides with an existing column."
            )

    new_columns = [
        {"name": renames.get(c["name"], c["name"]), "type": c["type"]}
        for c in columns
    ]
    return new_columns, rows


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------


def _read_csv_header(content: bytes) -> list[str]:
    """Cheaply read ONLY the CSV header row (first line) for the width pre-check.

    ``csv.reader`` over a ``StringIO`` parses lazily, so ``next()`` consumes just
    the first line — we learn the column count WITHOUT the O(cols) per-column
    inference or the O(cols^2) sanitize that ``infer_schema_streaming`` would do
    on a pathological wide header.
    """
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise UploadParseError("File is not valid UTF-8 CSV.") from exc
    reader = csv.reader(io.StringIO(text))
    return next(reader, [])


def _parse_csv(
    content: bytes, max_rows: int | None
) -> tuple[list[dict[str, str]], list[list[Any]]]:
    """Parse CSV bytes into ({name,type} columns, string-cell rows).

    Uses ``infer_schema_streaming`` for the header + per-column types + total row
    count (a whole-file scan so nullability/widening is detected everywhere), then
    a second ``csv.reader`` pass for the data rows.  The input is byte-bounded by
    the endpoint's pre-parse cap, so the full-row read is safe.
    """
    # (HIGH #2) Cap header WIDTH before any O(cols)/O(cols^2) work below.
    if len(_read_csv_header(content)) > MAX_UPLOAD_COLUMNS:
        raise UploadParseError(
            f"Upload has too many columns (max {MAX_UPLOAD_COLUMNS})."
        )

    try:
        header, inferred_types, total = infer_schema_streaming(content)
    except ValueError as exc:
        raise UploadParseError("File is not valid UTF-8 CSV.") from exc

    if not header:
        raise UploadParseError("File has no header row.")

    if max_rows is not None and total > max_rows:
        raise ScratchTooLargeError(
            f"Upload has {total} rows, over the {max_rows}-row scratch cap."
        )

    sanitized = _sanitize_columns(header)
    columns = [
        {"name": sanitized[i], "type": inferred_types[i]} for i in range(len(header))
    ]

    # Full data-row pass — cells stay strings.  (content already decoded cleanly
    # above via infer_schema_streaming, so this decode cannot fail.)
    text = content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows_iter = iter(reader)
    next(rows_iter, None)  # skip the header row

    n = len(header)
    rows: list[list[Any]] = [
        [row[i] if i < len(row) else "" for i in range(n)] for row in rows_iter
    ]
    return columns, rows


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------


def _stringify(cell: Any) -> str:
    """Stringify one XLSX cell so it converges onto the CSV string-cell path.

    ``None`` (a blank cell) becomes ``""`` so it is treated exactly like a blank
    CSV cell (drives Nullable widening + coerces to NULL at insert).
    """
    return "" if cell is None else str(cell)


def _parse_xlsx(
    content: bytes, max_rows: int | None
) -> tuple[list[dict[str, str]], list[list[Any]]]:
    """Parse XLSX bytes into ({name,type} columns, string-cell rows).

    Streaming + hardened: ``read_only=True`` (row streaming, no whole-sheet load),
    ``data_only=True`` (cached values, no formula evaluation).  Every cell is
    stringified so the XLSX path reuses the CSV type-inference + coercion path.

    The row cap is enforced INSIDE the streaming loop — the check runs before a
    row is appended, so an over-cap sheet raises after at most ``max_rows`` rows
    are held in memory (never after materializing the whole sheet).
    """
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - dependency is declared
        raise UploadParseError(
            "XLSX support is unavailable on this server."
        ) from exc

    try:
        workbook = openpyxl.load_workbook(
            io.BytesIO(content), read_only=True, data_only=True
        )
    except Exception as exc:  # noqa: BLE001 - any openpyxl failure = unreadable file
        raise UploadParseError("File is not a readable XLSX workbook.") from exc

    try:
        sheet = workbook.active
        if sheet is None:
            raise UploadParseError("XLSX workbook has no active sheet.")

        # (HIGH #2) Bound the READ width: openpyxl read_only pads every streamed
        # row out to the sheet's declared max_col (up to 16,384).  Capping max_col
        # here means a tiny file with an inflated dimension can never expand a row
        # past MAX_UPLOAD_COLUMNS + 1 cells in memory, regardless of the row cap.
        rows_iter = sheet.iter_rows(
            values_only=True, max_col=MAX_UPLOAD_COLUMNS + 1
        )
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise UploadParseError("File has no header row.")

        # openpyxl pads the header out to max_col too; strip trailing empty padding
        # to recover the REAL column count before enforcing the width cap.
        header_cells = list(header_row)
        while header_cells and header_cells[-1] is None:
            header_cells.pop()
        if len(header_cells) > MAX_UPLOAD_COLUMNS:
            raise UploadParseError(
                f"Upload has too many columns (max {MAX_UPLOAD_COLUMNS})."
            )

        header = [_stringify(c).strip() for c in header_cells]
        if not header or all(h == "" for h in header):
            raise UploadParseError("File has no header row.")

        n = len(header)
        rows: list[list[Any]] = []
        for raw in rows_iter:
            # Enforce the cap AS ROWS STREAM IN — before appending, so an over-cap
            # sheet never fully materializes in memory (mitigation #4).
            if max_rows is not None and len(rows) >= max_rows:
                raise ScratchTooLargeError(
                    f"Upload exceeds the {max_rows}-row scratch cap."
                )
            row = [_stringify(raw[i]) if i < len(raw) else "" for i in range(n)]
            rows.append(row)
    finally:
        workbook.close()

    # Per-column type inference over the stringified cells — the SAME rules as
    # the CSV path (infer_schema_streaming), so a fixture .xlsx and its equivalent
    # .csv produce an identical {columns, rows} shape (round-trip parity).
    sanitized = _sanitize_columns(header)
    columns: list[dict[str, str]] = []
    for i, name in enumerate(sanitized):
        col_values = [row[i] for row in rows]
        columns.append({"name": name, "type": infer_column_type(col_values)})
    return columns, rows
