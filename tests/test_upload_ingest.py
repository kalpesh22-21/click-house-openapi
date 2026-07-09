"""Unit + endpoint tests for the external-dataset upload front door (UI Slice 4).

Pure unit tests — NO live ClickHouse.  They prove the load-bearing behaviours:

  * ``parse_upload`` CSV → {name,type} columns + string-cell rows, with sanitized
    headers, BOM stripping, and Nullable widening on a late empty cell;
  * ``parse_upload`` XLSX → the SAME {columns, rows} shape as the equivalent CSV
    (round-trip parity), via an in-memory fixture .xlsx built with openpyxl;
  * the XLSX row cap is enforced *as rows stream in* — an over-cap sheet raises
    after at most ``max_rows`` rows are consumed, NEVER after draining the sheet;
  * ``apply_mapping`` renames role-tagged columns to canonical keys and rejects a
    duplicate role / a rename-target collision (fail-closed);
  * the endpoints: /scratch/v1/analyze previews (no CH); byte cap → 413
    UPLOAD_TOO_LARGE; row cap → 413 SCRATCH_TOO_LARGE; missing session → 400.

The authoritative /scratch/v1/upload route's real materialize needs a live
ClickHouse — that path is covered by an integration test
(tests/integration/test_upload_ingest_integration.py), skip-guarded.
"""

from __future__ import annotations

import io
import os

import pytest

_BASE_ENV = {
    "CLICKHOUSE_HOST": "localhost",
    "CLICKHOUSE_PORT": "9000",
    "CLICKHOUSE_USER": "default",
    "CLICKHOUSE_PASSWORD": "test-password",
    "CLICKHOUSE_DATABASE": "default",
    "CLICKHOUSE_SECURE": "false",
    "OIDC_JWKS_URL": "https://idp.test/.well-known/jwks.json",
    "OIDC_ISSUER": "https://idp.test/",
    "OIDC_AUDIENCE": "clickhouse-api",
    "JWT_ALGORITHMS": "RS256",
    "ALLOWED_DATABASES": "*",
    "APP_PORT": "8000",
    "LOG_LEVEL": "INFO",
    "PUBLIC_BASE_URL": "https://test.example.com",
}
for _k, _v in _BASE_ENV.items():
    os.environ.setdefault(_k, _v)

import openpyxl  # noqa: E402

from app import upload_ingest  # noqa: E402
from app.scratch_ingest import ScratchTooLargeError  # noqa: E402
from app.upload_ingest import (  # noqa: E402
    MAX_UPLOAD_COLUMNS,
    UploadMappingError,
    UploadParseError,
    apply_mapping,
    parse_upload,
)


# ---------------------------------------------------------------------------
# Fixtures / helpers
# ---------------------------------------------------------------------------


def _xlsx_bytes(rows: list[list[object]]) -> bytes:
    """Build an in-memory .xlsx whose active sheet holds *rows* (first = header)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ===========================================================================
# A. CSV parsing
# ===========================================================================


class TestParseCSV:
    def test_columns_types_and_string_rows(self):
        csv_bytes = b"Emp ID,Dept Name,FTE\n1001,Engineering,1.0\n1002,Sales,0.5\n"
        columns, rows = parse_upload(csv_bytes, "people.csv")
        assert columns == [
            {"name": "emp_id", "type": "Int64"},
            {"name": "dept_name", "type": "String"},
            {"name": "fte", "type": "Float64"},
        ]
        # Cells stay STRINGS (coerced to the declared type only at insert).
        assert rows == [["1001", "Engineering", "1.0"], ["1002", "Sales", "0.5"]]

    def test_nullable_widening_on_late_empty_cell(self):
        # A blank cell anywhere in an otherwise-int column widens it to Nullable.
        csv_bytes = b"code\n1\n2\n\n4\n"
        columns, rows = parse_upload(csv_bytes, "x.csv")
        assert columns == [{"name": "code", "type": "Nullable(Int64)"}]
        assert rows == [["1"], ["2"], [""], ["4"]]

    def test_bom_is_stripped_from_first_header(self):
        csv_bytes = "﻿Name,Value\nA,1\n".encode("utf-8")
        columns, _ = parse_upload(csv_bytes, "bom.csv")
        assert columns[0]["name"] == "name"  # no BOM char smuggled into the id

    def test_duplicate_headers_are_disambiguated(self):
        csv_bytes = b"id,id,ID\n1,2,3\n"
        columns, _ = parse_upload(csv_bytes, "dup.csv")
        assert [c["name"] for c in columns] == ["id", "id_2", "id_3"]

    def test_short_and_long_rows_normalized_to_column_count(self):
        csv_bytes = b"a,b,c\n1,2\n4,5,6,7\n"
        _, rows = parse_upload(csv_bytes, "ragged.csv")
        assert rows == [["1", "2", ""], ["4", "5", "6"]]  # padded / truncated to 3

    def test_empty_file_raises_parse_error(self):
        with pytest.raises(UploadParseError):
            parse_upload(b"", "empty.csv")

    def test_non_utf8_raises_parse_error(self):
        with pytest.raises(UploadParseError):
            parse_upload(b"\xff\xfebad\x00bytes", "bad.csv")

    def test_row_cap_exceeded_raises_too_large(self):
        csv_bytes = b"k\n" + b"\n".join(str(i).encode() for i in range(5)) + b"\n"
        with pytest.raises(ScratchTooLargeError) as e:
            parse_upload(csv_bytes, "big.csv", max_rows=3)
        assert e.value.code == "SCRATCH_TOO_LARGE"
        assert e.value.status_code == 413

    def test_row_cap_exactly_at_limit_allowed(self):
        csv_bytes = b"k\n1\n2\n3\n"
        _, rows = parse_upload(csv_bytes, "atcap.csv", max_rows=3)
        assert len(rows) == 3

    def test_too_many_columns_rejected(self):
        # A header wider than the cap (here also all-duplicate, the O(cols^2)
        # sanitize vector) → UPLOAD_PARSE_ERROR before any per-column work.
        header = ",".join(["a"] * (MAX_UPLOAD_COLUMNS + 1)).encode()
        csv_bytes = header + b"\n" + b",".join([b"1"] * (MAX_UPLOAD_COLUMNS + 1)) + b"\n"
        with pytest.raises(UploadParseError):
            parse_upload(csv_bytes, "wide.csv")

    def test_exactly_at_column_cap_allowed(self):
        header = ",".join([f"c{i}" for i in range(MAX_UPLOAD_COLUMNS)]).encode()
        csv_bytes = header + b"\n"
        cols, _ = parse_upload(csv_bytes, "atcolcap.csv")
        assert len(cols) == MAX_UPLOAD_COLUMNS


# ===========================================================================
# B. XLSX parsing — round-trip parity with CSV + streaming cap
# ===========================================================================


class TestParseXLSX:
    def test_xlsx_matches_equivalent_csv_shape(self):
        # Same logical data as CSV; integers + strings + a nullable-int column so
        # the stringified XLSX cells match the CSV text exactly (no float ambiguity).
        data = [
            ["Emp ID", "Dept Name", "Level"],
            [1001, "Engineering", 5],
            [1002, "Sales", None],  # blank → Nullable(Int64) widening
        ]
        xlsx = _xlsx_bytes(data)
        x_cols, x_rows = parse_upload(xlsx, "people.xlsx")

        csv_bytes = b"Emp ID,Dept Name,Level\n1001,Engineering,5\n1002,Sales,\n"
        c_cols, c_rows = parse_upload(csv_bytes, "people.csv")

        assert x_cols == c_cols
        assert x_rows == c_rows
        assert x_cols == [
            {"name": "emp_id", "type": "Int64"},
            {"name": "dept_name", "type": "String"},
            {"name": "level", "type": "Nullable(Int64)"},
        ]

    def test_xlsx_headers_sanitized(self):
        xlsx = _xlsx_bytes([["Order Date", "Total (USD)"], ["2024-01-01", "5"]])
        cols, _ = parse_upload(xlsx, "orders.xlsx")
        assert [c["name"] for c in cols] == ["order_date", "total_usd"]

    def test_unreadable_xlsx_raises_parse_error(self):
        with pytest.raises(UploadParseError):
            parse_upload(b"not a zip / not xlsx", "broken.xlsx")

    def test_xlsx_too_many_real_columns_rejected(self):
        # A header wider than the cap → UPLOAD_PARSE_ERROR at the header stage,
        # before any data row is streamed or inferred.
        wide_header = [f"c{i}" for i in range(MAX_UPLOAD_COLUMNS + 1)]
        xlsx = _xlsx_bytes([wide_header, [1] * len(wide_header)])
        with pytest.raises(UploadParseError):
            parse_upload(xlsx, "wide.xlsx")

    def test_xlsx_inflated_dimension_is_read_bounded_not_blown_up(self):
        # A "sideways zip bomb": a few real header columns but a far cell inflates
        # the sheet dimension to column 16,384.  The max_col read bound means those
        # far cells are never materialized — parse SUCCEEDS with just the real cols.
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["id", "name", "note"])
        ws.append([1, "a", "x"])
        ws.append([2, "b", "y"])
        ws.cell(row=2, column=16384, value="FAR")  # inflate the dimension
        buf = io.BytesIO()
        wb.save(buf)
        cols, rows = parse_upload(buf.getvalue(), "inflated.xlsx")
        assert [c["name"] for c in cols] == ["id", "name", "note"]  # 3, not 16384
        assert all(len(r) == 3 for r in rows)

    def test_xlsx_formula_cell_not_evaluated_under_data_only(self):
        """A formula cell read under data_only=True must NEVER be evaluated and the
        formula text must NEVER leak into the parsed rows.  openpyxl returns the
        cached value (None for a workbook openpyxl itself wrote, since it does not
        compute formulas), which stringifies to "" — so the parser can neither
        execute '=A2*2' nor surface the literal '=A2*2'."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["a", "b"])  # header
        ws["A2"] = 5
        ws["B2"] = "=A2*2"  # a live formula in the data row
        buf = io.BytesIO()
        wb.save(buf)

        columns, rows = parse_upload(buf.getvalue(), "formula.xlsx")

        assert [c["name"] for c in columns] == ["a", "b"]
        # The formula cell stringifies to "" (cached value is None) — NOT evaluated
        # to "10" and NOT leaked as the literal "=A2*2".
        assert rows == [["5", ""]]
        assert "=A2*2" not in {cell for r in rows for cell in r}
        assert "10" not in {cell for r in rows for cell in r}

    def test_xlsx_row_cap_streams_early_does_not_drain_sheet(self, monkeypatch):
        """The over-cap XLSX check runs AS ROWS STREAM — it must raise after at most
        max_rows rows are consumed, NOT after draining a (huge) sheet.

        A lazy counting sheet stands in for openpyxl: it would yield ~1e9 rows if
        fully drained, but the parser must stop early.  We assert it consumed only
        header + (max_rows + 1) rows before raising SCRATCH_TOO_LARGE."""
        max_rows = 5
        huge = 1_000_000_000

        class _CountingSheet:
            def __init__(self) -> None:
                self.consumed = 0

            def iter_rows(self, values_only=True, max_col=None):  # noqa: ANN001
                self.consumed += 1
                yield ("id", "name")  # header
                for i in range(huge):
                    self.consumed += 1
                    yield (i, f"name{i}")

        class _FakeWB:
            def __init__(self, sheet) -> None:  # noqa: ANN001
                self.active = sheet

            def close(self) -> None:
                pass

        sheet = _CountingSheet()
        import openpyxl as _oxl

        # load_workbook ignores its (BytesIO) arg and returns the lazy fake sheet.
        monkeypatch.setattr(_oxl, "load_workbook", lambda *a, **k: _FakeWB(sheet))

        with pytest.raises(ScratchTooLargeError) as e:
            parse_upload(b"ignored-bytes", "huge.xlsx", max_rows=max_rows)
        assert e.value.code == "SCRATCH_TOO_LARGE"
        # header (1) + max_rows appended + 1 that tripped the cap = max_rows + 2.
        assert sheet.consumed == max_rows + 2
        assert sheet.consumed < huge  # proved: the sheet was NOT drained


# ===========================================================================
# C. apply_mapping — rename + fail-closed validation
# ===========================================================================


def _cols(*names: str) -> list[dict[str, str]]:
    return [{"name": n, "type": "String"} for n in names]


class TestApplyMapping:
    def test_renames_roles_to_canonical_keys(self):
        columns = _cols("emp", "dept", "dept_label", "note")
        rows = [["e1", "d1", "Sales", "hi"]]
        mapping = {
            "emp": "EmployeeCode",
            "dept": "DepartmentCode",
            "dept_label": "DepartmentName",
            "note": "none",
        }
        new_cols, new_rows = apply_mapping(columns, rows, mapping)
        assert [c["name"] for c in new_cols] == [
            "employee_code",
            "department_code",
            "department_name",
            "note",
        ]
        assert new_rows is rows  # rows untouched (rename is columns-only)

    def test_absent_and_none_columns_keep_their_name(self):
        columns = _cols("a", "b")
        new_cols, _ = apply_mapping(columns, [], {"a": "none"})
        assert [c["name"] for c in new_cols] == ["a", "b"]

    def test_unknown_mapping_key_is_ignored(self):
        columns = _cols("a")
        new_cols, _ = apply_mapping(columns, [], {"ghost": "EmployeeCode"})
        assert [c["name"] for c in new_cols] == ["a"]

    def test_duplicate_role_rejected(self):
        columns = _cols("a", "b")
        with pytest.raises(UploadMappingError):
            apply_mapping(columns, [], {"a": "EmployeeCode", "b": "EmployeeCode"})

    def test_rename_target_collides_with_kept_column_rejected(self):
        # A column already named 'employee_code' is KEPT while another is mapped to
        # EmployeeCode → the canonical target collides → reject fail-closed.
        columns = _cols("foo", "employee_code")
        with pytest.raises(UploadMappingError):
            apply_mapping(columns, [], {"foo": "EmployeeCode"})

    def test_unknown_role_value_rejected(self):
        with pytest.raises(UploadMappingError):
            apply_mapping(_cols("a"), [], {"a": "Manager"})

    def test_non_dict_mapping_rejected(self):
        with pytest.raises(UploadMappingError):
            apply_mapping(_cols("a"), [], ["not", "a", "dict"])  # type: ignore[arg-type]

    def test_mapping_own_name_to_its_role_is_noop_not_collision(self):
        # A column sanitized to 'employee_code' AND mapped to EmployeeCode is a
        # no-op rename (target == its own name), not a collision.
        columns = _cols("employee_code", "other")
        new_cols, _ = apply_mapping(columns, [], {"employee_code": "EmployeeCode"})
        assert [c["name"] for c in new_cols] == ["employee_code", "other"]


# ===========================================================================
# D. Endpoints — analyze preview, byte cap, row cap, missing session
# ===========================================================================

from starlette.testclient import TestClient  # noqa: E402

from app.config import get_settings  # noqa: E402
from app.mcp_server import JWTAuthMiddleware, mcp  # noqa: E402
from tests.jwt_helpers import make_jwt  # noqa: E402


def _authed_app():
    return JWTAuthMiddleware(mcp.streamable_http_app(), settings=get_settings())


def _hdrs(session_id: str | None):
    h = {"Authorization": f"Bearer {make_jwt(session_id=session_id)}"}
    if session_id is not None:
        h["X-Session-Id"] = session_id
    return h


class TestAnalyzeRoute:
    def test_analyze_returns_preview(self):
        client = TestClient(_authed_app(), raise_server_exceptions=False)
        csv_bytes = b"Emp ID,Dept\n1001,Eng\n1002,Sales\n"
        resp = client.post(
            "/scratch/v1/analyze",
            headers=_hdrs("sessowner"),
            files={"file": ("people.csv", csv_bytes, "text/csv")},
        )
        assert resp.status_code == 200
        body = resp.json()
        assert body["columns"] == [
            {"name": "emp_id", "type": "Int64"},
            {"name": "dept", "type": "String"},
        ]
        assert body["row_count"] == 2
        assert body["sample_rows"] == [["1001", "Eng"], ["1002", "Sales"]]

    def test_analyze_missing_session_400(self):
        client = TestClient(_authed_app(), raise_server_exceptions=False)
        resp = client.post(
            "/scratch/v1/analyze",
            headers={"Authorization": f"Bearer {make_jwt()}"},
            files={"file": ("x.csv", b"a\n1\n", "text/csv")},
        )
        assert resp.status_code == 400
        assert resp.json()["code"] == "SCRATCH_SESSION_MISSING"

    def test_analyze_byte_cap_413(self, monkeypatch):
        monkeypatch.setenv("UPLOAD_MAX_BYTES", "10")
        get_settings.cache_clear()
        try:
            client = TestClient(_authed_app(), raise_server_exceptions=False)
            resp = client.post(
                "/scratch/v1/analyze",
                headers=_hdrs("sessowner"),
                files={"file": ("big.csv", b"a,b,c\n1,2,3\n4,5,6\n", "text/csv")},
            )
        finally:
            get_settings.cache_clear()
        assert resp.status_code == 413
        assert resp.json()["code"] == "UPLOAD_TOO_LARGE"

    def test_analyze_row_cap_413(self, monkeypatch):
        monkeypatch.setenv("SCRATCH_MAX_ROWS", "2")
        get_settings.cache_clear()
        try:
            client = TestClient(_authed_app(), raise_server_exceptions=False)
            resp = client.post(
                "/scratch/v1/analyze",
                headers=_hdrs("sessowner"),
                files={"file": ("big.csv", b"k\n1\n2\n3\n4\n", "text/csv")},
            )
        finally:
            get_settings.cache_clear()
        assert resp.status_code == 413
        assert resp.json()["code"] == "SCRATCH_TOO_LARGE"

    def test_analyze_parse_error_400(self):
        client = TestClient(_authed_app(), raise_server_exceptions=False)
        resp = client.post(
            "/scratch/v1/analyze",
            headers=_hdrs("sessowner"),
            files={"file": ("empty.csv", b"", "text/csv")},
        )
        assert resp.status_code == 400
        assert resp.json()["code"] == "UPLOAD_PARSE_ERROR"

    def test_analyze_corrupt_xlsx_parse_error_400(self):
        # A file with a .xlsx extension whose bytes are not a readable workbook must
        # surface as 400 UPLOAD_PARSE_ERROR at the endpoint — never a 500 crash.
        client = TestClient(_authed_app(), raise_server_exceptions=False)
        resp = client.post(
            "/scratch/v1/analyze",
            headers=_hdrs("sessowner"),
            files={
                "file": (
                    "broken.xlsx",
                    b"PK\x03\x04 not really a workbook",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert resp.status_code == 400
        assert resp.json()["code"] == "UPLOAD_PARSE_ERROR"

    def test_analyze_missing_file_part_400(self):
        client = TestClient(_authed_app(), raise_server_exceptions=False)
        resp = client.post(
            "/scratch/v1/analyze",
            headers=_hdrs("sessowner"),
            data={"notfile": "x"},
        )
        assert resp.status_code == 400
        assert resp.json()["code"] == "UPLOAD_PARSE_ERROR"


class TestUploadRouteFrontDoor:
    """Upload-route paths that fail BEFORE materialize (no live CH needed)."""

    def test_upload_missing_session_400(self):
        client = TestClient(_authed_app(), raise_server_exceptions=False)
        resp = client.post(
            "/scratch/v1/upload",
            headers={"Authorization": f"Bearer {make_jwt()}"},
            files={"file": ("x.csv", b"a\n1\n", "text/csv")},
            data={"mapping": "{}"},
        )
        assert resp.status_code == 400
        assert resp.json()["code"] == "SCRATCH_SESSION_MISSING"

    def test_upload_malformed_mapping_400(self):
        client = TestClient(_authed_app(), raise_server_exceptions=False)
        resp = client.post(
            "/scratch/v1/upload",
            headers=_hdrs("sessowner"),
            files={"file": ("x.csv", b"a\n1\n", "text/csv")},
            data={"mapping": "{not json"},
        )
        assert resp.status_code == 400
        assert resp.json()["code"] == "UPLOAD_MAPPING_INVALID"

    def test_upload_duplicate_role_mapping_400(self):
        client = TestClient(_authed_app(), raise_server_exceptions=False)
        resp = client.post(
            "/scratch/v1/upload",
            headers=_hdrs("sessowner"),
            files={"file": ("x.csv", b"a,b\n1,2\n", "text/csv")},
            data={"mapping": '{"a": "EmployeeCode", "b": "EmployeeCode"}'},
        )
        assert resp.status_code == 400
        assert resp.json()["code"] == "UPLOAD_MAPPING_INVALID"

    def test_upload_byte_cap_413(self, monkeypatch):
        monkeypatch.setenv("UPLOAD_MAX_BYTES", "5")
        get_settings.cache_clear()
        try:
            client = TestClient(_authed_app(), raise_server_exceptions=False)
            resp = client.post(
                "/scratch/v1/upload",
                headers=_hdrs("sessowner"),
                files={"file": ("x.csv", b"a,b,c\n1,2,3\n", "text/csv")},
                data={"mapping": "{}"},
            )
        finally:
            get_settings.cache_clear()
        assert resp.status_code == 413
        assert resp.json()["code"] == "UPLOAD_TOO_LARGE"


class TestResourceExhaustionHardening:
    """HIGH resource-exhaustion blockers from the Slice-4 review."""

    def test_content_length_over_cap_rejected_before_parse(self, monkeypatch):
        """A body whose declared Content-Length exceeds the cap (+ overhead) is 413
        BEFORE the parser runs — the server never buffers/parses the oversized body."""
        import app.mcp_server as mcp_server

        def _boom(*a, **k):  # parse_upload must never be reached
            raise AssertionError("parser was invoked on an over-cap body")

        monkeypatch.setattr(mcp_server, "parse_upload", _boom)
        monkeypatch.setenv("UPLOAD_MAX_BYTES", "10")
        get_settings.cache_clear()
        try:
            client = TestClient(_authed_app(), raise_server_exceptions=False)
            # ~130 KB body → Content-Length far exceeds 10 + 64 KB overhead.
            big = b"a\n" + b"x" * 130_000
            resp = client.post(
                "/scratch/v1/analyze",
                headers=_hdrs("sessowner"),
                files={"file": ("big.csv", big, "text/csv")},
            )
        finally:
            get_settings.cache_clear()
        assert resp.status_code == 413
        assert resp.json()["code"] == "UPLOAD_TOO_LARGE"

    def test_oversized_mapping_field_rejected(self):
        """The non-file `mapping` part (buffered fully in RAM by Starlette) is bounded
        by the fixed 64 KiB max_part_size — a 512-col mapping is ≤ ~30 KiB, so an
        over-64 KiB field is 413.  Uses the default 8 MB cap so the total body stays
        under the Content-Length pre-check and the max_part_size layer is exercised."""
        client = TestClient(_authed_app(), raise_server_exceptions=False)
        resp = client.post(
            "/scratch/v1/upload",
            headers=_hdrs("sessowner"),
            files={"file": ("x.csv", b"a\n1\n", "text/csv")},
            data={"mapping": "a" * (70 * 1024)},  # > 64 KiB part cap, < 8 MB CL slack
        )
        assert resp.status_code == 413
        assert resp.json()["code"] == "UPLOAD_TOO_LARGE"

    def test_wide_header_upload_rejected_400(self):
        client = TestClient(_authed_app(), raise_server_exceptions=False)
        header = ",".join(["a"] * (MAX_UPLOAD_COLUMNS + 1)).encode()
        resp = client.post(
            "/scratch/v1/analyze",
            headers=_hdrs("sessowner"),
            files={"file": ("wide.csv", header + b"\n1\n", "text/csv")},
        )
        assert resp.status_code == 400
        assert resp.json()["code"] == "UPLOAD_PARSE_ERROR"

    def test_too_many_form_fields_rejected_413(self):
        """Part-COUNT guard (max_fields): a body with more than file + mapping is
        refused — ~999 buffered text fields can never accumulate in FormData."""
        client = TestClient(_authed_app(), raise_server_exceptions=False)
        extra = {f"f{i}": "x" for i in range(10)}  # 10 tiny text fields
        resp = client.post(
            "/scratch/v1/upload",
            headers=_hdrs("sessowner"),
            files={"file": ("x.csv", b"a\n1\n", "text/csv")},
            data=extra,
        )
        assert resp.status_code == 413
        assert resp.json()["code"] == "UPLOAD_TOO_LARGE"


# ---------------------------------------------------------------------------
# Bounded receive-wrapper — the total-bytes backstop (RAM + disk).
# Drives _read_upload_part directly with a CHUNKED, no-Content-Length body so
# the CL pre-check can't fire and only the wrapper / form limits catch it.
# ---------------------------------------------------------------------------

import asyncio  # noqa: E402

import app.mcp_server as mcp_server  # noqa: E402
from app.upload_ingest import UploadTooLargeError  # noqa: E402
from starlette.requests import Request as _StarletteRequest  # noqa: E402

_BOUNDARY = "BOUNDX"


def _multipart_body(file_bytes: bytes, fields: dict[str, str] | None = None) -> bytes:
    parts = [
        (
            f"--{_BOUNDARY}\r\n"
            'Content-Disposition: form-data; name="file"; filename="big.csv"\r\n'
            "Content-Type: text/csv\r\n\r\n"
        ).encode()
        + file_bytes
        + b"\r\n"
    ]
    for k, v in (fields or {}).items():
        parts.append(
            (f"--{_BOUNDARY}\r\n" f'Content-Disposition: form-data; name="{k}"\r\n\r\n').encode()
            + v.encode()
            + b"\r\n"
        )
    parts.append(f"--{_BOUNDARY}--\r\n".encode())
    return b"".join(parts)


async def _read_chunked(body: bytes, max_bytes: int, chunk: int = 8192):
    """Feed *body* to _read_upload_part in chunks with NO Content-Length header.

    Returns (result_or_exc, chunks_consumed) so a test can assert the wrapper
    tripped BEFORE the whole body was drained.
    """
    chunks = [body[i : i + chunk] for i in range(0, len(body), chunk)]
    state = {"consumed": 0}
    it = iter(chunks)

    async def receive():
        try:
            b = next(it)
            state["consumed"] += 1
            return {"type": "http.request", "body": b, "more_body": True}
        except StopIteration:
            return {"type": "http.request", "body": b"", "more_body": False}

    scope = {
        "type": "http",
        "method": "POST",
        "path": "/scratch/v1/upload",
        # NOTE: no content-length header → the pre-check cannot fire.
        "headers": [
            (b"content-type", f"multipart/form-data; boundary={_BOUNDARY}".encode())
        ],
    }
    req = _StarletteRequest(scope, receive)
    try:
        result = await mcp_server._read_upload_part(req, max_bytes)
        return ("ok", result), state["consumed"], len(chunks)
    except BaseException as exc:  # noqa: BLE001
        return ("raise", exc), state["consumed"], len(chunks)


class TestBoundedReceiveWrapper:
    def test_cumulative_file_bytes_trip_before_draining(self):
        # 200 KB file, cap 10 (+64 KB slack) → the wrapper must raise mid-stream,
        # long before the full 200 KB is drained to disk.
        body = _multipart_body(b"x" * 200_000, {"mapping": "{}"})
        (kind, exc), consumed, total = asyncio.run(_read_chunked(body, max_bytes=10))
        assert kind == "raise"
        assert isinstance(exc, UploadTooLargeError)
        # Tripped early: far fewer chunks consumed than the whole body.
        assert consumed < total
        # ~ (10 + 64 KiB) / 8 KiB ≈ 9 chunks, nowhere near the ~25 total.
        assert consumed <= total // 2

    def test_no_content_length_small_body_happy_path(self):
        # A normal small CSV with no Content-Length flows through both the wrapper
        # and the form limits untouched.
        body = _multipart_body(b"a,b\n1,2\n", {"mapping": "{}"})
        (kind, result), consumed, total = asyncio.run(_read_chunked(body, max_bytes=8 * 1024 * 1024))
        assert kind == "ok"
        content, filename, form = result
        assert content == b"a,b\n1,2\n"
        assert filename == "big.csv"
        assert form.get("mapping") == "{}"

    def test_too_many_fields_no_content_length_rejected(self):
        # Many tiny text fields (small cumulative bytes, so the wrapper passes) are
        # stopped by the max_fields part-count guard inside form().
        body = _multipart_body(b"a\n1\n", {f"f{i}": "x" for i in range(10)})
        (kind, exc), _, _ = asyncio.run(_read_chunked(body, max_bytes=8 * 1024 * 1024))
        assert kind == "raise"
        assert isinstance(exc, UploadTooLargeError)


class TestRoutesNotTools:
    def test_upload_routes_are_not_mcp_tools(self):
        import asyncio

        tools = asyncio.run(mcp.list_tools())
        names = {t.name for t in tools}
        assert not any(
            "upload" in n.lower() or "analyze" in n.lower() for n in names
        )
